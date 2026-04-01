from __future__ import annotations

import logging
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from config import AppConfig, CONFIG
from eval import EvalOutputs, QualityIssue

logger = logging.getLogger(__name__)


SHEETS_FIXED = [
    "cleaned_raw",
    "overview",
    "quality_issues",
    "accuracy_breakdown",
    "confusion_matrix",
    "top_wrong_cases",
    "keyword_risk",
    "slice_analysis",
    "suggestions",
]


def _auto_fit_columns(ws, df: pd.DataFrame, max_width: int = 80) -> None:
    # 简单列宽优化（不追求完美）
    for i, col in enumerate(df.columns, start=1):
        series = df[col].astype(str).fillna("")
        max_len = max([len(str(col))] + [len(x) for x in series.head(200).tolist()])
        width = min(max(10, max_len + 2), max_width)
        ws.column_dimensions[get_column_letter(i)].width = width


def write_df(ws, df: pd.DataFrame, start_row: int, start_col: int, title: Optional[str] = None) -> int:
    """
    写入 df 到 openpyxl worksheet，返回写完后的下一行（空一行）
    """
    r = start_row
    c = start_col

    if title:
        ws.cell(row=r, column=c, value=title)
        r += 1

    # header
    for j, col in enumerate(df.columns, start=c):
        ws.cell(row=r, column=j, value=str(col))
    r += 1

    # body
    for _, row in df.iterrows():
        for j, col in enumerate(df.columns, start=c):
            val = row[col]
            ws.cell(row=r, column=j, value=_safe_cell_value(val))
        r += 1

    return r + 1  # 空一行


def _safe_cell_value(v: Any) -> Any:
    if pd.isna(v):
        return ""
    # 避免写入 numpy 类型导致 openpyxl 不兼容
    try:
        if hasattr(v, "item"):
            return v.item()
    except Exception:
        pass
    return v


def export_report(outputs: EvalOutputs, cfg: AppConfig = CONFIG) -> None:
    out_path = cfg.paths.output_excel

    wb = Workbook()
    # 删除默认 Sheet
    default = wb.active
    wb.remove(default)

    # 创建固定 sheet
    ws_map = {name: wb.create_sheet(title=name) for name in SHEETS_FIXED}

    # 1) cleaned_raw：先用 openpyxl 手动写（便于后续按坐标上色）
    cleaned = outputs.cleaned_raw_df.copy()
    ws_clean = ws_map["cleaned_raw"]
    _write_cleaned_raw(ws_clean, cleaned)
    _apply_cleaned_styles(ws_clean, cleaned, outputs.issues_to_style, cfg)

    # 2) overview：多表堆叠
    ws_ov = ws_map["overview"]
    r = 1
    for key, df in outputs.overview_tables.items():
        r = write_df(ws_ov, df, start_row=r, start_col=1, title=key)

    # 3) quality_issues：summary + reason_distribution
    ws_q = ws_map["quality_issues"]
    quality = outputs.meta.get("quality_issues", {})
    r = 1
    if isinstance(quality, dict) and "summary" in quality:
        r = write_df(ws_q, quality["summary"], start_row=r, start_col=1, title="summary")
    if isinstance(quality, dict) and "reason_distribution" in quality:
        r = write_df(ws_q, quality["reason_distribution"], start_row=r, start_col=1, title="reason_distribution")

    # 4) accuracy_breakdown
    ws_acc = ws_map["accuracy_breakdown"]
    write_df(ws_acc, outputs.accuracy_breakdown_df, start_row=1, start_col=1, title="accuracy_breakdown")

    # 5) confusion_matrix：两种模式
    ws_cm = ws_map["confusion_matrix"]
    _write_confusion_sheet(ws_cm, outputs.confusion_payload)

    # 6) top_wrong_cases
    ws_tw = ws_map["top_wrong_cases"]
    write_df(ws_tw, outputs.top_wrong_cases_df, start_row=1, start_col=1, title="top_wrong_cases")

    # 7) keyword_risk：两张表
    ws_kw = ws_map["keyword_risk"]
    r = 1
    for key, df in outputs.keyword_risk_tables.items():
        r = write_df(ws_kw, df, start_row=r, start_col=1, title=key)

    # 8) slice_analysis：多张表堆叠
    ws_sl = ws_map["slice_analysis"]
    r = 1
    for key, df in outputs.slice_tables.items():
        r = write_df(ws_sl, df, start_row=r, start_col=1, title=key)

    # 9) suggestions
    ws_sg = ws_map["suggestions"]
    write_df(ws_sg, outputs.suggestions_df, start_row=1, start_col=1, title="suggestions")

    # 基础列宽优化（cleaned_raw 已单独做）
    for name, ws in ws_map.items():
        if name == "cleaned_raw":
            continue
        # 找到第一个 table（粗略按 A2 作为 header 位置）
        # 这里不强求完美，避免开销过大
        pass

    wb.save(out_path)
    logger.info("已输出报告：%s", out_path)


def _write_cleaned_raw(ws, df: pd.DataFrame) -> None:
    # header
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=str(col))
    # body
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=i, column=j, value=_safe_cell_value(row[col]))

    _auto_fit_columns(ws, df, max_width=90)


def _apply_cleaned_styles(ws, df: pd.DataFrame, issues: List[QualityIssue], cfg: AppConfig) -> None:
    invalid_fill = PatternFill("solid", fgColor=cfg.style.invalid_fill_hex.replace("#", ""))
    dup_fill = PatternFill("solid", fgColor=cfg.style.duplicate_id_fill_hex.replace("#", ""))

    col_to_idx = {col: idx for idx, col in enumerate(df.columns, start=1)}

    for it in issues:
        if it.col_name not in col_to_idx:
            continue
        r = it.excel_row_1based
        c = col_to_idx[it.col_name]
        cell = ws.cell(row=r, column=c)
        if it.issue_type == "invalid":
            cell.fill = invalid_fill
        elif it.issue_type == "duplicate_id":
            cell.fill = dup_fill


def _write_confusion_sheet(ws, payload: Dict[str, Any]) -> None:
    mode = payload.get("mode", "unavailable")
    r = 1
    ws.cell(row=r, column=1, value="mode")
    ws.cell(row=r, column=2, value=str(mode))
    r += 2

    note = payload.get("note", "")
    if note:
        ws.cell(row=r, column=1, value="note")
        ws.cell(row=r, column=2, value=str(note))
        r += 2

    if mode == "available":
        cm = payload.get("confusion_matrix")
        metrics = payload.get("metrics")
        rows_used = payload.get("rows_used_for_confusion")
        dropped = payload.get("rows_dropped_invalid_human_decision")

        ws.cell(row=r, column=1, value="rows_used_for_confusion")
        ws.cell(row=r, column=2, value=rows_used if rows_used is not None else "")
        r += 1
        ws.cell(row=r, column=1, value="rows_dropped_invalid_human_decision")
        ws.cell(row=r, column=2, value=dropped if dropped is not None else "")
        r += 2

        if isinstance(cm, pd.DataFrame):
            r = write_df(ws, cm, start_row=r, start_col=1, title="confusion_matrix: human_decision x decision")
        if isinstance(metrics, pd.DataFrame):
            r = write_df(ws, metrics, start_row=r, start_col=1, title="metrics: false_delete_rate / missed_delete_rate")

    else:
        tf = payload.get("tf_by_decision")
        if isinstance(tf, pd.DataFrame):
            r = write_df(ws, tf, start_row=r, start_col=1, title="fallback: human_check(TRUE/FALSE) x decision")
